import os
import pandas as pd
from datetime import datetime
from data_collector import fetch_live_account, fetch_market_data
from quant_engine import load_state, calculate_v4_params
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "stock_portfolio_log.xlsx")

def log_portfolio_to_excel():
    print("📊 [고급 엑셀 퀀트 대시보드 로깅 및 시각화 시스템] 가동...")
    
    account_data = fetch_live_account()
    state = load_state()
    
    total_asset_krw = 0
    cash_usd = 0
    tot_eval_usd = 0
    tot_profit_usd = 0
    tot_pft_rt = 0

    if account_data and "Output_0" in account_data:
        summary = account_data["Output_0"]
        total_asset_krw = float(summary.get("tot_aet_amt", 0))
        cash_usd = float(summary.get("fc_aet_amt", 0))
        tot_eval_usd = float(summary.get("fc_eal_amt", 0))
        tot_profit_usd = float(summary.get("fc_eal_pls_amt", 0))
        tot_pft_rt = float(summary.get("pft_rt", 0))

    today_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    soxl_state = state.get("SOXL", {"cycle": 1, "T": 0.0})
    tqqq_state = state.get("TQQQ", {"cycle": 1, "T": 0.0})
    
    summary_data = [{
        "기록일시": today_date,
        "총평가자산(원)": total_asset_krw,
        "예수금($)": cash_usd,
        "주식평가금액($)": tot_eval_usd,
        "평가손익($)": tot_profit_usd,
        "총수익률(%)": tot_pft_rt,
        "SOXL 사이클": soxl_state['cycle'],
        "SOXL T회차": soxl_state['T'],
        "TQQQ 사이클": tqqq_state['cycle'],
        "TQQQ T회차": tqqq_state['T']
    }]
    
    df_summary = pd.DataFrame(summary_data)
    
    holdings_list = []
    if account_data and "Output_1" in account_data:
        for h in account_data["Output_1"]:
            code = h.get("iem_cd")
            qty = float(h.get("cns_bse_bnc_qty", 0))
            avg_p = float(h.get("fc_phs_uit_pr", 0))
            cur_p = float(h.get("fc_sec_end_pr", 0))
            pft = float(h.get("eal_pft_rt", 0))
            
            holdings_list.append({
                "기록일시": today_date,
                "티커": code,
                "종목명": h.get("iem_nm"),
                "보유수량": qty,
                "평균단가($)": avg_p,
                "현재가($)": cur_p,
                "평가손익률(%)": pft
            })
            
    df_holdings = pd.DataFrame(holdings_list)
    
    sheet_summary = "자산요약대시보드"
    sheet_holdings = "보유종목상세"
    
    if os.path.exists(EXCEL_FILE):
        try:
            existing_summary = pd.read_excel(EXCEL_FILE, sheet_name=sheet_summary)
            updated_summary = pd.concat([existing_summary, df_summary], ignore_index=True)
        except Exception:
            updated_summary = df_summary
            
        try:
            existing_holdings = pd.read_excel(EXCEL_FILE, sheet_name=sheet_holdings)
            updated_holdings = pd.concat([existing_holdings, df_holdings], ignore_index=True)
        except Exception:
            updated_holdings = df_holdings
    else:
        updated_summary = df_summary
        updated_holdings = df_holdings
        
    try:
        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
            updated_summary.to_excel(writer, sheet_name=sheet_summary, index=False)
            if not updated_holdings.empty:
                updated_holdings.to_excel(writer, sheet_name=sheet_holdings, index=False)
                
        # Professional Styling & Chart Dashboard via openpyxl
        wb = openpyxl.load_workbook(EXCEL_FILE)
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
        cell_font = Font(name="맑은 고딕", size=10)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        for s_name in [sheet_summary, sheet_holdings]:
            if s_name in wb.sheetnames:
                ws = wb[s_name]
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                for row in range(2, ws.max_row + 1):
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.font = cell_font
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                
                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

        if sheet_summary in wb.sheetnames:
            ws = wb[sheet_summary]
            ws._charts.clear()
            max_row = ws.max_row
            
            if max_row > 1:
                chart1 = LineChart()
                chart1.title = "📈 포트폴리오 총평가자산(원) 추이"
                chart1.style = 10
                chart1.y_axis.title = "총평가자산 (KRW)"
                chart1.x_axis.title = "기록일시"
                chart1.width = 20
                chart1.height = 12
                
                data1 = Reference(ws, min_col=2, min_row=1, max_row=max_row)
                cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
                chart1.add_data(data1, titles_from_data=True)
                chart1.set_categories(cats)
                ws.add_chart(chart1, "L2")
                
                chart2 = LineChart()
                chart2.title = "🎯 총수익률(%) 및 현금 추이"
                chart2.style = 13
                chart2.y_axis.title = "수익률 (%)"
                chart2.x_axis.title = "기록일시"
                chart2.width = 20
                chart2.height = 12
                
                data2 = Reference(ws, min_col=6, min_row=1, max_row=max_row)
                chart2.add_data(data2, titles_from_data=True)
                chart2.set_categories(cats)
                ws.add_chart(chart2, "L18")

                chart3 = LineChart()
                chart3.title = "⚡ 종목별 T회차 진행 상황"
                chart3.style = 2
                chart3.y_axis.title = "T회차 (0~40)"
                chart3.x_axis.title = "기록일시"
                chart3.width = 20
                chart3.height = 12
                
                data3 = Reference(ws, min_col=8, min_row=1, max_row=max_row)
                data3_2 = Reference(ws, min_col=10, min_row=1, max_row=max_row)
                chart3.add_data(data3, titles_from_data=True)
                chart3.add_data(data3_2, titles_from_data=True)
                chart3.set_categories(cats)
                ws.add_chart(chart3, "L34")

        wb.save(EXCEL_FILE)
        print(f"✅ 고급 엑셀 퀀트 대시보드 로깅 및 다중 차트 생성 완료: {EXCEL_FILE}")
    except PermissionError:
        print(f"⚠️ [엑셀 저장 실패] '{EXCEL_FILE}' 파일이 현재 열려있어 닫은 후 다시 실행해야 갱신됩니다.")
    except Exception as e:
        print(f"⚠️ 엑셀 처리 중 오류 발생: {e}")

if __name__ == "__main__":
    log_portfolio_to_excel()
