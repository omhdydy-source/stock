import os
import json
import urllib.request
import urllib.parse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from vr_engine import calculate_vr_cycle
from config import TELEGRAM_TOKEN, CHAT_ID

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_LOG_PATH = os.path.join(BASE_DIR, "vr_portfolio_log.xlsx")

def update_excel_log(rep):
    target_path = EXCEL_LOG_PATH
    if os.path.exists(target_path):
        try:
            wb = openpyxl.load_workbook(target_path)
            if "루나시트 VR일지" in wb.sheetnames:
                ws = wb["루나시트 VR일지"]
            else:
                ws = wb.active
                ws.title = "루나시트 VR일지"
        except Exception:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "루나시트 VR일지"
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "루나시트 VR일지"

    navy_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    light_blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    font_title = Font(name="맑은 고딕", size=14, bold=True, color="1F4E78")
    font_header = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="맑은 고딕", size=10, bold=True)
    font_normal = Font(name="맑은 고딕", size=10)
    
    thin = Side(border_style="thin", color="D3D3D3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.cell(row=1, column=2, value="🚀 라오어식 VR 5.0 오피셜 루나시트 대시보드").font = font_title
    ws.row_dimensions[1].height = 30

    summary_headers = ["총 자산 ($)", "주식 평가금 ($)", "예수금 Pool ($)", "기준 목표선 V ($)", "안전 밴드 하단", "안전 밴드 상단", "진단 액션"]
    summary_vals = [
        rep["total_asset"],
        rep["total_stock_eval"],
        rep["current_Pool"],
        rep["current_V"],
        rep["v_min"],
        rep["v_max"],
        rep["action"]
    ]

    for c_idx, h_text in enumerate(summary_headers, 2):
        c_cell = ws.cell(row=3, column=c_idx, value=h_text)
        c_cell.fill = navy_fill
        c_cell.font = font_header
        c_cell.alignment = Alignment(horizontal="center", vertical="center")
        c_cell.border = border

        v_cell = ws.cell(row=4, column=c_idx, value=summary_vals[c_idx-2])
        v_cell.fill = light_blue_fill
        v_cell.font = font_bold
        v_cell.alignment = Alignment(horizontal="right", vertical="center")
        v_cell.border = border
        if c_idx <= 8:
            v_cell.number_format = "$#,##0.00"
        else:
            v_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 24

    table_headers = [
        "회차", "일자", "TQQQ평단($)", "TQQQ보유수", "주식평가금($)", "Pool현금($)", 
        "총자산($)", "기준목표선V($)", "Pool비중(%)", "G값", "총상승률(%)", 
        "다음목표선(Next V)", "하단선(V min)", "상단선(V max)", "진단결과", "매매필요금액($)"
    ]

    start_row = 7
    for c_idx, h_text in enumerate(table_headers, 1):
        cell = ws.cell(row=start_row, column=c_idx, value=h_text)
        cell.fill = navy_fill
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[start_row].height = 28

    today_str = datetime.now().strftime("%Y-%m-%d")

    # Find existing rows and determine sequential cycle number (1회차, 2회차, 3회차...)
    data_row_idx = start_row + 1
    existing_rows_count = 0
    found_today = False

    while ws.cell(row=data_row_idx, column=1).value is not None and data_row_idx < 1000:
        existing_date = ws.cell(row=data_row_idx, column=2).value
        if existing_date == today_str:
            found_today = True
            break
        existing_rows_count += 1
        data_row_idx += 1

    if not found_today:
        cycle_num = existing_rows_count + 1
    else:
        cycle_num = ws.cell(row=data_row_idx, column=1).value or (existing_rows_count + 1)

    row_data = [
        cycle_num,
        today_str,
        round(rep["tqqq_avg_p"], 2),
        round(rep["tqqq_qty"], 2),
        round(rep["total_stock_eval"], 2),
        round(rep["current_Pool"], 2),
        round(rep["total_asset"], 2),
        round(rep["current_V"], 2),
        round(rep["pool_ratio"] / 100.0, 4),
        f"/{int(rep['G_value'])}",
        round(rep["total_rate"] / 100.0, 4),
        round(rep["next_V"], 2),
        round(rep["v_min"], 2),
        round(rep["v_max"], 2),
        rep["action"],
        round(rep["trade_amount"], 2)
    ]

    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=data_row_idx, column=col_idx, value=val)
        cell.border = border
        cell.font = font_normal
        
        if col_idx in [1, 2, 10, 15]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            
        if col_idx in [3, 5, 6, 7, 8, 12, 13, 14, 16]:
            if col_idx == 4:
                cell.number_format = "#,##0"
            else:
                cell.number_format = "$#,##0.00"
        elif col_idx in [9, 11]:
            cell.number_format = "0.00%"

        if col_idx == 15:
            cell.font = font_bold
            if val == "BUY":
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            elif val == "SELL":
                cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    ws.row_dimensions[data_row_idx].height = 22

    tier_start_row = data_row_idx + 3
    
    ws.cell(row=tier_start_row, column=1, value="🟢 [매수 30단계 차등가이드]").font = font_bold
    ws.cell(row=tier_start_row, column=9, value="🔴 [매도 30단계 차등가이드]").font = font_bold
    
    tier_headers = ["회차", "지정가($)", "할당금액($)", "수량(주)"]
    
    for idx, th in enumerate(tier_headers, 1):
        c = ws.cell(row=tier_start_row+1, column=idx, value=th)
        c.fill = navy_fill
        c.font = font_header
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    for idx, th in enumerate(tier_headers, 9):
        c = ws.cell(row=tier_start_row+1, column=idx, value=th)
        c.fill = navy_fill
        c.font = font_header
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    for i, t in enumerate(rep["buy_tier_orders"]):
        r = tier_start_row + 2 + i
        ws.cell(row=r, column=1, value=t["tier"]).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=t["price"]).number_format = "$#,##0.00"
        ws.cell(row=r, column=3, value=t["cost"]).number_format = "$#,##0.00"
        ws.cell(row=r, column=4, value=t["shares"]).number_format = "#,##0"
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = border
            ws.cell(row=r, column=c).font = font_normal

    for i, t in enumerate(rep["sell_tier_orders"]):
        r = tier_start_row + 2 + i
        ws.cell(row=r, column=9, value=t["tier"]).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=10, value=t["price"]).number_format = "$#,##0.00"
        ws.cell(row=r, column=11, value=t["revenue"]).number_format = "$#,##0.00"
        ws.cell(row=r, column=12, value=t["shares"]).number_format = "#,##0"
        for c in range(9, 13):
            ws.cell(row=r, column=c).border = border
            ws.cell(row=r, column=c).font = font_normal

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    ws._charts.clear()
    chart = LineChart()
    chart.title = "루나시트 VR 5.0 총자산 및 목표선 추이"
    chart.style = 13
    chart.y_axis.title = "금액 ($)"
    chart.x_axis.title = "일자"
    chart.width = 18
    chart.height = 12
    chart.legend.position = "t"
    chart.legend.include_in_layout = True

    data = Reference(ws, min_col=5, min_row=start_row, max_col=7, max_row=data_row_idx)
    cats = Reference(ws, min_col=2, min_row=start_row+1, max_row=data_row_idx)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "R7")

    try:
        wb.save(target_path)
        print(f"📁 구글시트 루나시트 오피셜 템플릿 양식 엑셀 누적 저장 완료: {target_path}")
    except PermissionError:
        alt_path = os.path.join(BASE_DIR, f"vr_portfolio_log_{datetime.now().strftime('%H%M%S')}.xlsx")
        wb.save(alt_path)
        print(f"⚠️ 기존 엑셀 파일이 열려 있어 대체 파일로 저장했습니다: {alt_path}")

def run_vr_bot():
    rep = calculate_vr_cycle()
    update_excel_log(rep)
    print("✅ 루나시트 양식 엑셀 파일 누적 업데이트 완료!")

if __name__ == "__main__":
    run_vr_bot()
